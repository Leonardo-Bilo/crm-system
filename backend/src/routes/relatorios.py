from flask import Blueprint, request, jsonify, send_file
from src.models.cliente import db, Cliente, HistoricoCompra, Agendamento, Lembrete
from datetime import datetime
from sqlalchemy import func
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
import io
import os
import tempfile
from src.models.cliente import StatusAgendamento

relatorios_bp = Blueprint('relatorios', __name__)

@relatorios_bp.route('/relatorios/exportar/clientes/excel', methods=['GET'])
def exportar_clientes_excel():
    try:
        # Criar workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Clientes"
        
        # Cabeçalhos
        headers = ['Nome', 'Email', 'Telefone', 'Endereço', 'Categoria', 'Notas', 'Data de Compra', 'Total de Compras', 'Valor Total']
        
        # Estilo do cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Buscar clientes
        clientes = Cliente.query.filter_by(ativo=True).all()
        
        # Preencher dados
        for row, cliente in enumerate(clientes, 2):
            ws.cell(row=row, column=1, value=cliente.nome)
            ws.cell(row=row, column=2, value=cliente.email)
            ws.cell(row=row, column=3, value=cliente.telefone or "")
            ws.cell(row=row, column=4, value=cliente.endereco or "")
            ws.cell(row=row, column=5, value=cliente.categoria.nome if cliente.categoria else "")
            ws.cell(row=row, column=6, value=cliente.notas or "")
            # Buscar data da última compra
            if cliente.historico_compras:
                ultima_compra = max(cliente.historico_compras, key=lambda x: x.data_compra)
                data_ultima_compra = ultima_compra.data_compra.strftime("%d/%m/%Y") if ultima_compra.data_compra else ""
            else:
                data_ultima_compra = ""
            ws.cell(row=row, column=7, value=data_ultima_compra)
            ws.cell(row=row, column=8, value=sum([compra.quantidade for compra in cliente.historico_compras]))
            ws.cell(row=row, column=9, value=sum([compra.valor for compra in cliente.historico_compras]))
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Salvar em arquivo temporário
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name=f'clientes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'erro': str(e)}), 500

@relatorios_bp.route('/relatorios/exportar/clientes/pdf', methods=['GET'])
def exportar_clientes_pdf():
    try:
        # Criar arquivo temporário
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
        
        # Criar documento PDF
        doc = SimpleDocTemplate(temp_file.name, pagesize=A4)
        story = []
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center
        )
        
        # Título
        title = Paragraph("Relatório de Clientes", title_style)
        story.append(title)
        story.append(Spacer(1, 12))
        
        # Data do relatório
        data_relatorio = Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}", styles['Normal'])
        story.append(data_relatorio)
        story.append(Spacer(1, 20))
        
        # Buscar clientes
        clientes = Cliente.query.filter_by(ativo=True).all()
        
        # Dados da tabela
        data = [['Nome', 'Email', 'Telefone', 'Categoria', 'Total de Compras', 'Valor Total']]
        
        for cliente in clientes:
            data.append([
                cliente.nome,
                cliente.email,
                cliente.telefone or "",
                cliente.categoria.nome if cliente.categoria else "",
                sum([compra.quantidade for compra in cliente.historico_compras]),
                f"R$ {sum([compra.valor for compra in cliente.historico_compras]):.2f}"
            ])
        
        # Criar tabela
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        
        # Construir PDF
        doc.build(story)
        temp_file.close()
        
        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name=f'clientes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf',
            mimetype='application/pdf'
        )
        
    except Exception as e:
        return jsonify({'erro': str(e)}), 500

@relatorios_bp.route('/relatorios/exportar/vendas/excel', methods=['GET'])
def exportar_vendas_excel():
    try:
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        
        # Criar workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Vendas"
        
        # Cabeçalhos
        headers = ['Data', 'Cliente', 'Produto/Serviço', 'Quantidade', 'Valor Unitário', 'Valor Total', 'Status']
        
        # Estilo do cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Buscar vendas
        query = HistoricoCompra.query.join(Cliente)
        
        if data_inicio:
            query = query.filter(HistoricoCompra.data_compra >= datetime.fromisoformat(data_inicio))
        if data_fim:
            query = query.filter(HistoricoCompra.data_compra <= datetime.fromisoformat(data_fim))
        
        vendas = query.all()
        
        # Preencher dados
        for row, venda in enumerate(vendas, 2):
            ws.cell(row=row, column=1, value=venda.data_compra.strftime("%d/%m/%Y"))
            ws.cell(row=row, column=2, value=venda.cliente.nome)
            ws.cell(row=row, column=3, value=venda.produto_servico)
            ws.cell(row=row, column=4, value=venda.quantidade)
            ws.cell(row=row, column=5, value=venda.valor / venda.quantidade)
            ws.cell(row=row, column=6, value=venda.valor)
            ws.cell(row=row, column=7, value=venda.status)
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Salvar em arquivo temporário
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name=f'vendas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'erro': str(e)}), 500

@relatorios_bp.route('/relatorios/exportar/agendamentos/excel', methods=['GET'])
def exportar_agendamentos_excel():
    try:
        # Criar workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Agendamentos"
        
        # Cabeçalhos
        headers = ['Data', 'Cliente', 'Título', 'Tipo', 'Status', 'Local', 'Observações']
        
        # Estilo do cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Buscar agendamentos
        agendamentos = Agendamento.query.join(Cliente).all()
        
        # Preencher dados
        for row, agendamento in enumerate(agendamentos, 2):
            ws.cell(row=row, column=1, value=agendamento.data_agendamento.strftime("%d/%m/%Y %H:%M"))
            ws.cell(row=row, column=2, value=agendamento.cliente.nome)
            ws.cell(row=row, column=3, value=agendamento.titulo)
            ws.cell(row=row, column=4, value=agendamento.tipo or "")
            ws.cell(row=row, column=5, value=agendamento.status.value if agendamento.status else "")
            ws.cell(row=row, column=6, value=agendamento.local or "")
            ws.cell(row=row, column=7, value=agendamento.observacoes or "")
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Salvar em arquivo temporário
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name=f'agendamentos_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'erro': str(e)}), 500

# Rotas para buscar dados dos relatórios
@relatorios_bp.route('/relatorios/vendas', methods=['GET'])
def get_vendas():
    try:
        vendas = HistoricoCompra.query.join(Cliente).all()
        dados = []
        
        for venda in vendas:
            dados.append({
                'id': venda.id,
                'data': venda.data_compra.isoformat() if venda.data_compra else None,
                'cliente': venda.cliente.nome if venda.cliente else '',
                'produto_servico': venda.produto_servico,
                'quantidade': venda.quantidade,
                'valor': float(venda.valor),
                'status': venda.status
            })
        
        return jsonify({'dados': dados})
    except Exception as e:
        return jsonify({'erro': str(e)}), 500

@relatorios_bp.route('/relatorios/agendamentos', methods=['GET'])
def get_agendamentos():
    try:
        agendamentos = Agendamento.query.join(Cliente).all()
        dados = []
        
        for agendamento in agendamentos:
            # Mapear status do enum para strings que o frontend espera
            status_mapping = {
                StatusAgendamento.AGENDADO: 'agendado',
                StatusAgendamento.CONFIRMADO: 'confirmado',
                StatusAgendamento.CONCLUIDO: 'concluido',
                StatusAgendamento.CANCELADO: 'cancelado',
                StatusAgendamento.REAGENDADO: 'reagendado'
            }
            
            dados.append({
                'id': agendamento.id,
                'data': agendamento.data_agendamento.isoformat() if agendamento.data_agendamento else None,
                'cliente': agendamento.cliente.nome if agendamento.cliente else '',
                'titulo': agendamento.titulo,
                'tipo': agendamento.tipo,
                'status': status_mapping.get(agendamento.status, 'agendado'),
                'local': agendamento.local,
                'observacoes': agendamento.observacoes
            })
        
        return jsonify({'dados': dados})
    except Exception as e:
        return jsonify({'erro': str(e)}), 500

@relatorios_bp.route('/relatorios/lembretes', methods=['GET'])
def get_lembretes():
    try:
        lembretes = Lembrete.query.all()
        dados = []
        
        for lembrete in lembretes:
            dados.append({
                'id': lembrete.id,
                'data': lembrete.data_lembrete.isoformat() if lembrete.data_lembrete else None,
                'titulo': lembrete.titulo,
                'descricao': lembrete.descricao,
                'prioridade': lembrete.prioridade.value if lembrete.prioridade else 'media',
                'status': 'pendente' if not lembrete.concluido else 'concluido',
                'concluido': lembrete.concluido
            })
        
        return jsonify({'dados': dados})
    except Exception as e:
        return jsonify({'erro': str(e)}), 500

@relatorios_bp.route('/relatorios/resumo', methods=['GET'])
def get_resumo():
    try:
        # Total de clientes ativos
        total_clientes = Cliente.query.filter_by(ativo=True).count()
        
        # Total de vendas e valor
        vendas = HistoricoCompra.query.all()
        total_compras = len(vendas)
        valor_total_vendas = sum([float(v.valor) for v in vendas])
        
        # Agendamentos por status
        agendamentos_pendentes = Agendamento.query.filter_by(status=StatusAgendamento.AGENDADO).count()
        agendamentos_confirmados = Agendamento.query.filter_by(status=StatusAgendamento.CONFIRMADO).count()
        agendamentos_concluidos = Agendamento.query.filter_by(status=StatusAgendamento.CONCLUIDO).count()
        agendamentos_cancelados = Agendamento.query.filter_by(status=StatusAgendamento.CANCELADO).count()
        total_agendamentos = agendamentos_pendentes + agendamentos_confirmados + agendamentos_concluidos + agendamentos_cancelados
        
        # Lembretes pendentes
        lembretes_pendentes = Lembrete.query.filter_by(concluido=False).count()
        lembretes_concluidos = Lembrete.query.filter_by(concluido=True).count()
        total_lembretes = lembretes_pendentes + lembretes_concluidos
        
        return jsonify({
            'total_clientes': total_clientes,
            'total_compras': total_compras,
            'valor_total_vendas': valor_total_vendas,
            'agendamentos_pendentes': agendamentos_pendentes,
            'agendamentos_confirmados': agendamentos_confirmados,
            'agendamentos_concluidos': agendamentos_concluidos,
            'agendamentos_cancelados': agendamentos_cancelados,
            'total_agendamentos': total_agendamentos,
            'lembretes_pendentes': lembretes_pendentes,
            'lembretes_concluidos': lembretes_concluidos,
            'total_lembretes': total_lembretes
        })
    except Exception as e:
        return jsonify({'erro': str(e)}), 500

@relatorios_bp.route('/relatorios/vendas-periodo', methods=['GET'])
def get_vendas_periodo():
    try:
        data_inicio = request.args.get('data_inicio')
        data_fim = request.args.get('data_fim')
        
        query = HistoricoCompra.query
        
        if data_inicio:
            try:
                data_inicio_dt = datetime.fromisoformat(data_inicio)
            except Exception:
                data_inicio_dt = None
            if data_inicio_dt:
                query = query.filter(HistoricoCompra.data_compra >= data_inicio_dt)
        if data_fim:
            try:
                data_fim_dt = datetime.fromisoformat(data_fim)
            except Exception:
                data_fim_dt = None
            if data_fim_dt:
                query = query.filter(HistoricoCompra.data_compra <= data_fim_dt)
        vendas = query.all()
        # Agrupar por data
        vendas_por_data = {}
        for venda in vendas:
            data_obj = venda.data_compra
            # Se for string, tenta converter para datetime
            if isinstance(data_obj, str):
                try:
                    data_obj = datetime.fromisoformat(data_obj)
                except Exception:
                    data_obj = None
            # Garante que data_str seja sempre string
            if isinstance(data_obj, datetime):
                data_str = data_obj.strftime('%Y-%m-%d')
            elif isinstance(data_obj, str):
                data_str = data_obj  # já é string, usa direto
            else:
                data_str = 'sem-data'
            if data_str not in vendas_por_data:
                vendas_por_data[data_str] = {'total': 0, 'quantidade': 0, 'vendas': 0}
            vendas_por_data[data_str]['total'] += float(venda.valor)
            vendas_por_data[data_str]['quantidade'] += venda.quantidade
            vendas_por_data[data_str]['vendas'] += 1  # Conta cada transação como 1 venda
        dados = []
        for data, info in vendas_por_data.items():
            dados.append({
                'data': data,
                'total': info['total'],
                'quantidade': info['quantidade'],
                'vendas': info['vendas']
            })
        return jsonify(dados)
    except Exception as e:
        return jsonify({'erro': str(e)}), 500

